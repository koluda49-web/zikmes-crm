# -*- coding: utf-8 -*-
"""Генерация Excel-файла со списком телефонов для робота отправки СМС."""

from pathlib import Path

from openpyxl import Workbook, load_workbook


def write_sms_phones_file(phones: list, output_path: str) -> None:
    """
    Пишет телефоны в столбик, без заголовка и без лишних колонок —
    формат, аналогичный mts_numbers.csv из проекта Зикмес.
    Дубликаты телефонов (если один номер встретился в нескольких заказах)
    убираются, порядок сохраняется.
    """
    seen = set()
    unique_phones = []
    for phone in phones:
        if phone and phone not in seen:
            seen.add(phone)
            unique_phones.append(phone)

    wb = Workbook()
    ws = wb.active
    for phone in unique_phones:
        ws.append([phone])
    wb.save(output_path)


def append_sms_phones(phones: list, output_path: str) -> None:
    """
    Дозаписывает телефоны в конец файла (создаёт файл, если его ещё нет).
    Используется для инкрементального сохранения по ходу прогона, чтобы
    ручная остановка скрипта на середине не теряла уже найденные результаты.
    Дубликаты (в т.ч. уже присутствующие в файле) не добавляются повторно.
    """
    existing_phones = set()

    if Path(output_path).exists():
        wb = load_workbook(output_path)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            if row and row[0]:
                existing_phones.add(str(row[0]))
    else:
        wb = Workbook()
        ws = wb.active

    for phone in phones:
        if phone and phone not in existing_phones:
            ws.append([phone])
            existing_phones.add(phone)

    wb.save(output_path)

