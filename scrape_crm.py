#!/usr/bin/env python3
"""
Обзвон ДОЛЖНИКОВ через МТС VATS -> реестр должников (Google Таблица)

Список должников берётся из реестра (лист "Обзвон"), звонки идут через
отдельное задание в МТС с названием "Обзвон должников <дата время>",
результаты обзвона отправляются обратно в реестр (сопоставление по номеру
телефона делает Apps Script на стороне таблицы).

КОМАНДЫ:

1. Выгрузить номера должников из реестра и подготовить файл для МТС:
   python scrape_crm.py debtors_export

2. Создать в МТС задание "Обзвон должников <дата>" (свой текст сообщения,
   опрос "Итог"), загрузить номера, обзвон стартует сам:
   python scrape_crm.py call_debtors

3. Скачать отчёт по завершённому заданию и отправить результаты в реестр:
   python scrape_crm.py fetch_debtors_report [ССЫЛКА_НА_ЗАДАНИЕ]

4. (Вручную) Отправить уже скачанный отчёт МТС в реестр:
   python scrape_crm.py debtors_merge ОТЧЕТ_МТС.csv

5. ПОЛНЫЙ ЦИКЛ: пункты 1+2+3 подряд, в одном браузере:
   python scrape_crm.py debtors_full_cycle

Все файлы (номера, название задания, копия отчёта) складываются в папку output/
рядом со скриптом и перезаписываются при каждом запуске.

Настройка (в config.txt, рядом со скриптом):
    MTS_LOGIN=...
    MTS_PASSWORD=...
    DEBTOR_REGISTRY_WEBHOOK=<URL веб-приложения реестра должников>
    DEBTOR_REGISTRY_SECRET=<секрет, совпадающий с WEBHOOK_SECRET в registryScript.gs>

Установка:
    pip install selenium beautifulsoup4 webdriver-manager openpyxl pandas requests
"""

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
import platform

IS_CLOUD = 'RENDER' in os.environ or platform.system() != 'Windows'

# Все файлы, связанные с обзвоном должников, складываются сюда - чтобы не путать
# с файлами обычного обзвона по доставке (которые остаются рядом со скриптом,
# как и раньше) и было проще найти/каждый раз просто перезаписывается.
OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)
import csv
import time
import sys
import re
from datetime import datetime

import requests

CONFIG_FILE = "config.txt"


def load_config() -> dict:
    """Читает config.txt, затем переменные окружения (env перекрывают файл)."""
    config = {}
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, value = line.split("=", 1)
                config[key.strip()] = value.strip()
    for key in [
        'MTS_LOGIN', 'MTS_PASSWORD',
        'DEBTOR_REGISTRY_WEBHOOK', 'DEBTOR_REGISTRY_SECRET',
    ]:
        val = os.environ.get(key)
        if val:
            config[key] = val
    return config


def try_mts_login(driver):
    """
    Если открылась страница входа МТС — пытается автоматически ввести
    логин/пароль из config.txt. Если конфига нет или поля не найдены,
    просто ничего не делает (пользователь войдёт вручную как раньше).
    """
    config = load_config()
    login = config.get("MTS_LOGIN")
    password = config.get("MTS_PASSWORD")

    if not login or not password or password == "ВСТАВЬ_СЮДА_ПАРОЛЬ":
        return False  # конфиг не настроен — пропускаем автологин

    try:
        # Ищем поле логина: на форме входа МТС это первый input[type='text']
        # (без name='login'/'email' — просто обычное текстовое поле)
        login_field = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((
                By.CSS_SELECTOR,
                "input[type='text']"
            ))
        )
        password_field = driver.find_element(
            By.CSS_SELECTOR,
            "input[type='password']"
        )

        login_field.clear()
        login_field.send_keys(login)
        password_field.clear()
        password_field.send_keys(password)

        # Ищем кнопку входа
        try:
            submit_btn = driver.find_element(
                By.XPATH,
                "//button[@type='submit'] | //button[contains(.,'Войти')] | //button[contains(.,'Вход')]"
            )
            submit_btn.click()
        except Exception:
            from selenium.webdriver.common.keys import Keys
            password_field.send_keys(Keys.ENTER)

        print("   🔑 Логин выполнен автоматически (из config.txt)")
        time.sleep(2)
        return True

    except Exception:
        return False  # форма входа не найдена — значит уже авторизован или другая страница

LAST_TASK_URL_FILE  = "last_task_url.txt"   # ссылка на последнее созданное задание МТС

MTS_TASKS_LIST_URL = "https://193543133-18519208.vats.mts.by/#/autocaller/tasks"
MTS_NEW_TASK_URL = "https://193543133-18519208.vats.mts.by/#/autocaller/tasks/new"
MTS_SURVEY_NAME = "Итог"

# Текст для обзвона ДОЛЖНИКОВ.
MTS_MESSAGE_TEXT_DEBTORS = (
    "!!!-Здравствуйте! Это организация Зикмес! Вы просрочили платёж. "
    "Срочно оплатите или верните технику!"
)
DEBTOR_MTS_NUMBERS_FILE = os.path.join(OUTPUT_DIR, "mts_numbers_debtors.csv")  # отдельный файл номеров - не путать с обычным mts_numbers.csv
DEBTOR_TASK_NAME_PREFIX = "Обзвон должников"
DEBTOR_LAST_TASK_NAME_FILE = os.path.join(OUTPUT_DIR, "last_task_name_debtors.txt")  # своё название задания, не пересекается с обычным обзвоном

# В облаке ~/Downloads нет — скачиваем прямо в output/
DOWNLOADS_DIR = os.path.abspath(OUTPUT_DIR) if IS_CLOUD else os.path.join(os.path.expanduser("~"), "Downloads")

# URL и секрет веб-приложения РЕЕСТРА ДОЛЖНИКОВ читаются из config.txt (ключи
# DEBTOR_REGISTRY_WEBHOOK и DEBTOR_REGISTRY_SECRET) - впишите их туда после того,
# как развернёте doGet/doPost из registryScript.gs как веб-приложение. Ничего
# в этом файле для этого редактировать не нужно.
def get_debtor_registry_config():
    config = load_config()
    webhook = config.get("DEBTOR_REGISTRY_WEBHOOK", "")
    secret = config.get("DEBTOR_REGISTRY_SECRET", "")
    return webhook, secret


def fetch_debtor_batch() -> list[dict]:
    """
    Забирает из реестра должников список ещё не обработанных номеров
    (лист "Обзвон", где нет результата) через GET-запрос к веб-приложению.
    Возвращает список {id, fio, phone}.
    """
    webhook, secret = get_debtor_registry_config()
    if not webhook:
        print("  ⚠️  В config.txt не заполнен DEBTOR_REGISTRY_WEBHOOK - впишите туда URL веб-приложения реестра.")
        return []
    try:
        resp = requests.get(
            webhook,
            params={"action": "get_debtor_batch", "secret": secret},
            timeout=60,
        )
        result = resp.json()
        if not result.get("success"):
            print(f"  ⚠️  Ошибка получения списка должников: {result.get('message')}")
            return []
        return result.get("rows", [])
    except Exception as e:
        print(f"  ⚠️  Не удалось связаться с реестром должников: {e}")
        return []


def send_debtor_call_results(rows: list[dict]):
    """
    Отправляет результаты обзвона должников [{phone, result, date}] в реестр -
    Apps Script сам сопоставит по номеру телефона и запишет в лист "Обзвон".
    """
    webhook, secret = get_debtor_registry_config()
    if not webhook:
        print("  ⚠️  В config.txt не заполнен DEBTOR_REGISTRY_WEBHOOK - впишите туда URL веб-приложения реестра.")
        return
    try:
        resp = requests.post(
            webhook,
            json={
                "action": "upload_debtor_call_results",
                "secret": secret,
                "rows": rows,
            },
            timeout=60,
        )
        result = resp.json()
        if result.get("success"):
            print(f"  ☁️  Реестр должников обновлён: {result.get('message')}")
        else:
            print(f"  ⚠️  Ошибка записи в реестр должников: {result.get('message')}")
    except Exception as e:
        print(f"  ⚠️  Не удалось связаться с реестром должников: {e}")


# ============================================================
#  ЧАСТЬ 1 — Выгрузка из CRM
# ============================================================

def _find_pw_chromium():
    """Ищет Playwright-установленный Chromium и возвращает (путь, версия)."""
    import glob as _glob
    import subprocess as _sp
    pw_root = os.environ.get(
        'PLAYWRIGHT_BROWSERS_PATH', '/opt/render/project/src/.pw-browsers'
    )
    candidates = sorted(_glob.glob(
        os.path.join(pw_root, 'chromium-*/chrome-linux/chrome')
    ))
    if not candidates:
        return None, None
    binary = candidates[-1]
    try:
        out = _sp.check_output([binary, '--version'], stderr=_sp.DEVNULL, text=True)
        m = re.search(r'[\d]+\.[\d]+\.[\d]+\.[\d]+', out)
        return binary, (m.group(0) if m else None)
    except Exception:
        return binary, None


def create_driver():
    options = webdriver.ChromeOptions()
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)

    prefs = {
        "credentials_enable_service": False,
        "profile.password_manager_enabled": False,
        "profile.password_manager_leak_detection": False,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "download.default_directory": os.path.abspath(DOWNLOADS_DIR),
    }

    if IS_CLOUD:
        options.add_argument("--headless=new")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        options.add_argument("--disable-setuid-sandbox")
        options.add_argument("--window-size=1600,1000")

        # Используем Playwright-установленный Chromium как Chrome-бинарь для Selenium
        chrome_binary, chrome_version = _find_pw_chromium()
        if chrome_binary:
            print(f"   [MTS] Chromium: {chrome_binary} (v{chrome_version or '?'})")
            options.binary_location = chrome_binary
        else:
            print("   [MTS] ⚠ Playwright Chromium не найден, используем системный Chrome")
    else:
        profile_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "chrome_profile")
        os.makedirs(profile_dir, exist_ok=True)
        options.add_argument(f"--user-data-dir={profile_dir}")
        chrome_version = None

    options.add_experimental_option("prefs", prefs)

    try:
        from webdriver_manager.chrome import ChromeDriverManager
        if chrome_version:
            service = Service(ChromeDriverManager(driver_version=chrome_version).install())
        else:
            service = Service(ChromeDriverManager().install())
    except ImportError:
        service = Service()

    driver = webdriver.Chrome(service=service, options=options)
    if not IS_CLOUD:
        driver.maximize_window()
    return driver


# ============================================================
#  ЧАСТЬ 2 — Объединение отчёта МТС с заказами CRM
# ============================================================

def load_mts_report(filepath: str) -> dict:
    """
    Загружает отчёт МТС (CSV, новый .xlsx или старый .xls) в формате:
    Номер | Часовой пояс | Статус | Количество попыток | Результат последнего вызова |
    Длительность прослушивания | Аудиофайл | Ответ

    Возвращает {телефон: результат}
    """
    phone_to_result = {}
    lower = filepath.lower()

    if lower.endswith(".csv"):
        with open(filepath, "r", encoding="utf-8-sig") as f:
            sample = f.read(2048)
            f.seek(0)
            delimiter = ";" if sample.count(";") > sample.count(",") else ","
            reader = csv.DictReader(f, delimiter=delimiter)
            for row in reader:
                _store_mts_row(row, phone_to_result)
        return phone_to_result

    # Пытаемся открыть как нормальный .xlsx через openpyxl
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            row_dict = dict(zip(headers, row))
            _store_mts_row(row_dict, phone_to_result)
        return phone_to_result
    except Exception as e:
        print(f"  ⚠️  Не удалось открыть как .xlsx ({e})")
        print("  Пробую открыть как старый формат .xls / HTML-таблицу...")

    # Запасной вариант: файл на самом деле .xls (старый формат) или HTML-таблица
    # с расширением .xlsx — пробуем через pandas, который умеет и то и другое.
    import pandas as pd
    df = None
    for engine in (None, "xlrd", "html"):
        try:
            if engine == "html":
                tables = pd.read_html(filepath)
                df = tables[0]
            else:
                df = pd.read_excel(filepath, engine=engine) if engine else pd.read_excel(filepath)
            break
        except Exception:
            continue

    if df is None:
        raise RuntimeError(
            f"Не удалось прочитать файл отчёта '{filepath}' ни одним способом.\n"
            f"Попробуй скачать отчёт в формате CSV из МТС VATS вместо Excel."
        )

    df.columns = [str(c).strip() for c in df.columns]
    for _, row in df.iterrows():
        row_dict = row.to_dict()
        _store_mts_row(row_dict, phone_to_result)

    return phone_to_result


def _store_mts_row(row: dict, phone_to_result: dict):
    """Извлекает данные из одной строки отчёта МТС и сохраняет в словарь по номеру."""
    phone_raw = str(row.get("Номер", "")).strip()
    phone = normalize_phone(phone_raw)
    if not phone:
        return

    status = str(row.get("Статус", "")).strip()
    if status.lower() == "успешно":
        result_text = "Прослушано"
    elif status:
        result_text = f"Не дозвонились ({status})"
    else:
        result_text = "Нет данных"

    phone_to_result[phone] = result_text


def normalize_phone(raw: str) -> str:
    digits = re.sub(r'\D', '', raw)
    if digits.startswith("80"):
        digits = "375" + digits[2:]
    if digits.startswith("00375"):
        digits = digits[2:]
    if not digits.startswith("375"):
        return ""
    return digits[:12]



# ============================================================
#  ЧАСТЬ 2.5 — Обзвон ДОЛЖНИКОВ (реестр должников, отдельная таблица)
# ============================================================

def cmd_debtors_export():
    """
    Забирает из реестра должников список ещё не обзвоненных номеров
    и готовит файл mts_numbers.xlsx в ТОМ ЖЕ формате, что и раньше -
    поэтому дальше можно сразу запускать обычную команду 'call' без изменений.
    """
    print("📥 Забираю список должников из реестра...")
    debtors = fetch_debtor_batch()
    if not debtors:
        print("  Пусто - либо все уже обзвонены, либо реестр недоступен (проверь DEBTOR_REGISTRY_WEBHOOK в конфигурации).")
        return

    phones = set()
    for d in debtors:
        phone = normalize_phone(str(d.get("phone", "")))
        if re.fullmatch(r"375\d{9}", phone):
            phones.add(phone)
        else:
            print(f"  ⚠️  Пропущен некорректный номер (АЙДИ {d.get('id')}): '{d.get('phone')}'")

    if not phones:
        print("  ❌ Не нашлось ни одного корректного номера.")
        return

    mts_xlsx_file = DEBTOR_MTS_NUMBERS_FILE.replace(".csv", ".xlsx")
    wb_mts = openpyxl.Workbook()
    ws_mts = wb_mts.active
    ws_mts.title = "Номера"
    for i, phone in enumerate(sorted(phones), 1):
        cell = ws_mts.cell(row=i, column=1, value=int(phone))
        cell.number_format = "0"
    ws_mts.column_dimensions["A"].width = 18
    wb_mts.save(mts_xlsx_file)

    print(f"✅ Файл для МТС сохранён -> {mts_xlsx_file} ({len(phones)} уникальных номеров должников)")
    print("\nТеперь запусти: python scrape_crm.py call_debtors")


def cmd_call_new_debtors(external_driver=None):
    """
    Создаёт в МТС отдельное задание автообзвона ДОЛЖНИКОВ:
    название "Обзвон должников <дата время>", свой текст сообщения
    (MTS_MESSAGE_TEXT_DEBTORS), тот же опрос "Итог", свой файл номеров
    (DEBTOR_MTS_NUMBERS_FILE) - не пересекается с обычным обзвоном по доставке.
    """
    return cmd_call_new(
        external_driver=external_driver,
        task_name_prefix=DEBTOR_TASK_NAME_PREFIX,
        message_text=MTS_MESSAGE_TEXT_DEBTORS,
        numbers_file=DEBTOR_MTS_NUMBERS_FILE,
        task_name_file=DEBTOR_LAST_TASK_NAME_FILE,
    )


def cmd_debtors_merge(report_path: str):
    """
    Читает отчёт МТС по обзвону должников и отправляет результаты обратно
    в реестр (лист "Обзвон") - Apps Script сам сопоставит по номеру телефона.
    """
    print(f"📥 Загружаю отчёт МТС из {report_path}...")
    phone_to_result = load_mts_report(report_path)
    print(f"   Загружено результатов: {len(phone_to_result)}")

    rows = [
        {"phone": phone, "result": result, "date": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
        for phone, result in phone_to_result.items()
    ]

    print("\n☁️  Отправляю результаты обзвона в реестр должников...")
    send_debtor_call_results(rows)
    print("   Открой реестр должников - лист 'Обзвон' обновлён.")

    try:
        if os.path.exists(report_path):
            os.remove(report_path)
            print(f"   🗑️  Временный файл отчёта удалён: {os.path.basename(report_path)}")
    except Exception as e:
        print(f"   ⚠️  Не удалось удалить временный файл отчёта: {e}")


# ============================================================
#  ЧАСТЬ 3 — Автоматический запуск обзвона в МТС
# ============================================================

def cmd_call_new(external_driver=None, task_name_prefix=DEBTOR_TASK_NAME_PREFIX, message_text=None, numbers_file=None, task_name_file=None):
    """
    Команда: создать НОВОЕ чистое задание автоинформирования с нуля
    (вместо копирования старого), настроить синтез речи, сбор ответов,
    загрузить свежий файл номеров и сохранить.

    Это основной способ запуска обзвона — он не имеет проблемы с
    накоплением старых номеров, в отличие от копирования задания.

    task_name_prefix - префикс названия задания в МТС (дата добавится сама).
    message_text - текст для синтеза речи (по умолчанию MTS_MESSAGE_TEXT_DEBTORS).
    numbers_file - какой xlsx-файл с номерами загружать (по умолчанию DEBTOR_MTS_NUMBERS_FILE).
    """
    if message_text is None:
        message_text = MTS_MESSAGE_TEXT_DEBTORS
    if task_name_file is None:
        task_name_file = DEBTOR_LAST_TASK_NAME_FILE

    mts_xlsx_file = (numbers_file or DEBTOR_MTS_NUMBERS_FILE).replace(".csv", ".xlsx")
    if not os.path.exists(mts_xlsx_file):
        print(f"❌ Файл {mts_xlsx_file} не найден. Сначала запусти: python scrape_crm.py debtors_export")
        sys.exit(1)

    mts_xlsx_abspath = os.path.abspath(mts_xlsx_file)

    # Если передан внешний браузер (полный цикл) — используем его и не закрываем.
    # Если нет — создаём свой и закрываем в finally.
    own_driver = external_driver is None
    driver = external_driver if external_driver else create_driver()
    try:
        print("🌐 Открываю список заданий в МТС...")
        driver.get(MTS_TASKS_LIST_URL)
        time.sleep(2)

        logged_in_automatically = try_mts_login(driver)
        if logged_in_automatically:
            time.sleep(2)
            try:
                WebDriverWait(driver, 15).until(
                    lambda d: d.execute_script("return document.readyState") == "complete"
                )
            except Exception:
                pass
            if "tasks" not in driver.current_url:
                driver.get(MTS_TASKS_LIST_URL)
                time.sleep(2)

        print()
        print("=" * 55)
        if logged_in_automatically:
            print("  Вход выполнен автоматически.")
        else:
            print("  Если МТС попросит войти — авторизуйся в открывшемся окне.")
        print("  Дождись пока полностью загрузится список заданий.")
        if not IS_CLOUD:
            print("  Нажми Enter здесь, когда увидишь кнопку «Добавить задание».")
            print("=" * 55)
            input("  -> Нажми Enter: ")
        else:
            print("  Облачный режим: автоматическое ожидание кнопки «Добавить задание»...")
            print("=" * 55)
            # Ждём появления кнопки вместо input()
            try:
                WebDriverWait(driver, 30).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Добавить задание')]"))
                )
            except Exception:
                time.sleep(5)
        print()

        print("➕ Нажимаю «Добавить задание»...")
        add_btn = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Добавить задание')]"))
        )
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", add_btn)
        time.sleep(0.3)
        try:
            add_btn.click()
        except Exception:
            driver.execute_script("arguments[0].click();", add_btn)
        time.sleep(2)

        print("⏳ Жду загрузки формы создания задания...")
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.XPATH, "//*[contains(text(),'Основные данные')]"))
        )
        time.sleep(1)
        task_form_url = driver.current_url
        print(f"✅ Форма создания задания открыта: {task_form_url}")

        # 1. Название задания — с датой и временем, чтобы каждое было уникальным
        task_name = f"{task_name_prefix} {datetime.now().strftime('%d %m %Y %H %M')}"
        print(f"📝 Задаю название: {task_name}")
        # Сохраняем название сразу — пригодится при поиске строки в списке заданий
        try:
            with open(task_name_file, "w", encoding="utf-8") as f:
                f.write(task_name)
        except Exception:
            pass
        try:
            name_field = WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((
                    By.XPATH, "//label[contains(text(),'Название')]/following::input[1]"
                ))
            )
        except Exception:
            # Запасной вариант: первое текстовое поле на форме
            name_field = driver.find_element(By.CSS_SELECTOR, "input[type='text']")
        name_field.clear()
        name_field.send_keys(task_name)
        time.sleep(0.3)

        # 2. Выбираем "Синтез речи" для сообщения
        print("🔊 Выбираю «Синтез речи»...")
        synth_tab = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((
                By.XPATH, "//*[self::button or self::div or self::span][contains(text(),'Синтез речи')]"
            ))
        )
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", synth_tab)
        time.sleep(0.3)
        driver.execute_script("arguments[0].click();", synth_tab)
        time.sleep(1)

        # Проверяем что вкладка реально переключилась (должна появиться кнопка "Создать сообщение")
        try:
            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, "//button[contains(., 'Создать сообщение')]"))
            )
            print("   ✅ Вкладка «Синтез речи» активна")
        except Exception:
            print("   ⚠️  Похоже вкладка не переключилась, пробую кликнуть ещё раз...")
            driver.execute_script("arguments[0].click();", synth_tab)
            time.sleep(1)

        # 3. Нажимаем "Создать сообщение"
        create_msg_btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Создать сообщение')]"))
        )
        try:
            create_msg_btn.click()
        except Exception:
            driver.execute_script("arguments[0].click();", create_msg_btn)
        time.sleep(1)

        # 4. Вводим текст сообщения в открывшемся модальном окне
        print("✍️  Ввожу текст сообщения...")
        text_area = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//*[contains(text(),'Введите текст')]/following::textarea[1]"))
        )
        text_area.clear()
        text_area.send_keys(message_text)
        time.sleep(0.5)

        # 5. Нажимаем "Синтезировать речь"
        synth_btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Синтезировать речь')]"))
        )
        try:
            synth_btn.click()
        except Exception:
            driver.execute_script("arguments[0].click();", synth_btn)
        print("⏳ Жду синтеза речи (до 20 сек)...")
        time.sleep(5)

        # 6. Нажимаем "Применить"
        apply_btn = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Применить')]"))
        )
        try:
            apply_btn.click()
        except Exception:
            driver.execute_script("arguments[0].click();", apply_btn)
        print("   ✅ Сообщение применено")
        time.sleep(1)

        # 7. Загружаем файл номеров
        print(f"📤 Загружаю файл номеров: {mts_xlsx_abspath}")
        file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file'].file-input")
        file_input.send_keys(mts_xlsx_abspath)

        # 10. Подтверждаем модальное окно "Добавление номеров".
        # ВАЖНО: на странице ДВЕ кнопки с текстом "Загрузить список" —
        # одна обычная (открывает системный диалог выбора файла), и одна
        # внутри модального окна (подтверждает уже выбранный файл). Ищем
        # именно вторую — она находится рядом с текстом "будет добавлен".
        print("⏳ Жду модальное окно подтверждения номеров (опрашиваю каждые 0.5 сек, до 30 сек)...")
        confirmed = False
        deadline = time.time() + 30

        while time.time() < deadline and not confirmed:
            try:
                # Сначала находим сам текст модального окна — он появляется ТОЛЬКО
                # когда окно подтверждения открыто
                modal_marker = driver.find_element(By.XPATH, "//*[contains(text(),'будет добавлен')]")
                if modal_marker.is_displayed():
                    print(f"   Модальное окно найдено: «{modal_marker.text}»")

                    # Кнопка подтверждения находится в том же диалоговом контейнере,
                    # что и этот текст — ищем её через общего предка
                    confirm_btn = driver.find_element(
                        By.XPATH,
                        "//*[contains(text(),'будет добавлен')]/ancestor::*[self::div][3]//button[contains(., 'Загрузить список')]"
                    )

                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", confirm_btn)
                    time.sleep(0.3)
                    driver.execute_script("arguments[0].click();", confirm_btn)
                    time.sleep(1)

                    # Проверяем что модальное окно реально закрылось (текст исчез)
                    try:
                        still_there = driver.find_element(By.XPATH, "//*[contains(text(),'будет добавлен')]")
                        if still_there.is_displayed():
                            continue  # окно всё ещё открыто, клик не сработал — пробуем снова
                    except Exception:
                        pass  # текст пропал — окно закрылось, успех

                    confirmed = True
                    print("   ✅ Подтвердил загрузку списка номеров")
            except Exception:
                pass

            if not confirmed:
                time.sleep(0.5)

        if not confirmed:
            print("   ⚠️  Модальное окно подтверждения не появилось или не найдено.")
            if IS_CLOUD:
                print("   ⚠️  Облако: продолжаю без подтверждения — проверьте МТС вручную.")
            else:
                print("   ⚠️  ОБЯЗАТЕЛЬНО посмотри на экран браузера прямо сейчас:")
                print("   ⚠️  если видишь окно 'Добавление номеров' — нажми в нём")
                print("   ⚠️  кнопку 'Загрузить список' САМ, и только потом жми Enter здесь.")
                input("   -> Нажми Enter здесь ПОСЛЕ клика в браузере (или если окна не было): ")

        # 10.5 Настраиваем расписание: Пн-Пт, время 07:00 - 21:00
        print("📅 Настраиваю расписание (Пн-Пт, 07:00-21:00)...")
        try:
            schedule_link = driver.find_element(By.XPATH, "//*[contains(text(),'Настроить расписание')]")
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", schedule_link)
            time.sleep(0.3)
            driver.execute_script("arguments[0].click();", schedule_link)
            time.sleep(1)

            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//*[contains(text(),'Настройка расписания автоинформирования')]"))
            )

            # Находим модальное окно расписания целиком, чтобы искать поля
            # именно внутри него (а не по всей странице)
            modal = driver.find_element(
                By.XPATH, "//*[contains(text(),'Настройка расписания автоинформирования')]/ancestor::*[self::div][3]"
            )

            # Все текстовые поля времени внутри модального окна идут по порядку:
            # [Пн-с, Пн-по, Вт-с, Вт-по, Ср-с, Ср-по, Чт-с, Чт-по, Пт-с, Пт-по, ...]
            all_time_fields = modal.find_elements(By.CSS_SELECTOR, "input[type='text'], input:not([type])")
            print(f"   Найдено полей времени в окне: {len(all_time_fields)}")

            days = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница"]
            for i, day in enumerate(days):
                start_idx = i * 2
                if start_idx + 1 >= len(all_time_fields):
                    print(f"   ⚠️  Не хватает полей для «{day}»")
                    continue

                field_from = all_time_fields[start_idx]
                field_to = all_time_fields[start_idx + 1]

                for field, value in [(field_from, "07:00"), (field_to, "21:00")]:
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", field)
                    driver.execute_script("arguments[0].click();", field)
                    field.send_keys(Keys.CONTROL, "a")
                    field.send_keys(value)
                    field.send_keys(Keys.TAB)
                    time.sleep(0.2)
                print(f"   {day}: 07:00-21:00")

            # ВАЖНО: ищем кнопку "Сохранить" ТОЛЬКО внутри модального окна расписания
            # (переменная modal, найденная чуть выше), а не по всей странице. Раньше
            # поиск шёл по всему документу и брал "последнюю видимую" кнопку с текстом
            # "Сохранить" — но is_displayed() не учитывает перекрытие оверлеем, поэтому
            # под модалкой основная кнопка "Сохранить задание" тоже считалась "видимой",
            # и иногда клик улетал по НЕЙ — это сохраняло и закрывало всё задание сразу,
            # минуя настройку дозвона и включение информирования.
            schedule_save_btn = WebDriverWait(modal, 10).until(
                EC.presence_of_element_located((By.XPATH, ".//button[contains(., 'Сохранить')]"))
            )
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", schedule_save_btn)
            time.sleep(0.2)
            driver.execute_script("arguments[0].click();", schedule_save_btn)
            print("   ✅ Расписание настроено: Пн-Пт, 07:00-21:00")
            time.sleep(1)

            # Подстраховка: проверяем, что модальное окно расписания реально закрылось,
            # а не то, что страница перешла на сохранение всего задания (URL не должен
            # поменяться, и заголовок модалки должен исчезнуть)
            try:
                still_open = driver.find_elements(
                    By.XPATH, "//*[contains(text(),'Настройка расписания автоинформирования')]"
                )
                if any(el.is_displayed() for el in still_open):
                    print("   ⚠️  Модальное окно расписания всё ещё открыто — проверь вручную.")
            except Exception:
                pass

            # КРИТИЧЕСКАЯ ПРОВЕРКА: если после клика URL уехал от формы создания
            # задания — значит клик по ошибке попал по кнопке сохранения ВСЕГО
            # задания (а не модалки расписания), и задание уже сохранено/закрыто
            # без настройки дозвона и включения информирования. Продолжать
            # дальнейшие шаги на этой странице бессмысленно — лучше остановиться
            # сразу с понятной ошибкой, чем тихо наделать что-то ещё не туда.
            if driver.current_url != task_form_url:
                print("\n" + "🛑" * 20)
                print("  ОШИБКА: после сохранения расписания страница изменилась")
                print(f"  (было: {task_form_url}")
                print(f"   стало: {driver.current_url})")
                print("  Похоже, что вместо кнопки модалки расписания нажалась")
                print("  кнопка сохранения ВСЕГО задания, и оно уже сохранено")
                print("  БЕЗ настройки дозвона и БЕЗ включённого информирования!")
                print("  Зайди в МТС, найди это задание и донастрой его вручную:")
                print("   - Максимум попыток / перезвон через")
                print("   - Включить информирование")
                print("🛑" * 20)
                return
        except Exception as e:
            print(f"   ⚠️  Не удалось настроить расписание автоматически ({e}).")
            if IS_CLOUD:
                print("   ⚠️  Облако: продолжаю — донастройте расписание в МТС вручную.")
            else:
                print("   ⚠️  Настрой вручную: Пн-Пт, время 07:00-21:00, затем нажми Enter.")
                input("   -> Нажми Enter после настройки расписания вручную: ")

        # 10.6 Настройка дозвона: максимум попыток и перезвон
        print("\n📞 Настраиваю параметры дозвона...")
        if not IS_CLOUD:
            print("  ОСТАНОВКА: настрой параметры дозвона ВРУЧНУЮ на экране браузера:")
            print("  - Максимум попыток: выбери 3")
            print("  - Перезванивать через: впиши 5")
            input("  -> Когда настроишь — нажми Enter ЗДЕСЬ (не раньше!): ")
        else:
            # Пытаемся автоматизировать. Если не получится — задание создаётся
            # с дефолтными настройками; пользователь может подправить в МТС вручную.
            time.sleep(2)
            try:
                # Максимум попыток
                attempt_inputs = driver.find_elements(By.XPATH,
                    "//input[@type='number'] | //input[contains(@placeholder,'попыт')]"
                )
                for inp in attempt_inputs:
                    try:
                        if inp.is_displayed():
                            driver.execute_script("arguments[0].value='3'; arguments[0].dispatchEvent(new Event('input',{bubbles:true})); arguments[0].dispatchEvent(new Event('change',{bubbles:true}));", inp)
                            print("   ✅ Максимум попыток: 3")
                            break
                    except Exception:
                        continue
            except Exception as e:
                print(f"   ⚠️  Не удалось задать максимум попыток: {e}")

            try:
                # Перезвон через N минут
                interval_inputs = driver.find_elements(By.XPATH,
                    "//input[@type='number'] | //input[contains(@placeholder,'минут')]"
                )
                set_count = 0
                for inp in interval_inputs:
                    try:
                        if inp.is_displayed() and set_count == 1:
                            driver.execute_script("arguments[0].value='5'; arguments[0].dispatchEvent(new Event('input',{bubbles:true})); arguments[0].dispatchEvent(new Event('change',{bubbles:true}));", inp)
                            print("   ✅ Перезвон через: 5 мин")
                            break
                        if inp.is_displayed():
                            set_count += 1
                    except Exception:
                        continue
            except Exception as e:
                print(f"   ⚠️  Не удалось задать интервал перезвона: {e}")

        print("   Продолжаю...")
        time.sleep(0.5)

        # 11. "Включить информирование"
        print("\n🔘 Включаю информирование...")
        if not IS_CLOUD:
            print("  ОСТАНОВКА: включи тумблер «Включить информирование» ВРУЧНУЮ")
            print("  на экране браузера (должен стать синим/зелёным).")
            input("  -> Когда включишь — нажми Enter ЗДЕСЬ: ")
        else:
            time.sleep(1)
            try:
                # Ищем тумблер/переключатель "Включить информирование"
                toggles = driver.find_elements(By.XPATH,
                    "//*[contains(text(),'информирование')]/following::input[@type='checkbox'][1] | "
                    "//*[contains(text(),'информирование')]/following::*[contains(@class,'toggle') or contains(@class,'switch')][1]"
                )
                clicked = False
                for toggle in toggles:
                    try:
                        if toggle.is_displayed():
                            driver.execute_script("arguments[0].click();", toggle)
                            print("   ✅ Тумблер «Включить информирование» нажат")
                            clicked = True
                            break
                    except Exception:
                        continue
                if not clicked:
                    # Запасной вариант: ищем label с текстом
                    labels = driver.find_elements(By.XPATH, "//label[contains(.,'информирование')]")
                    for label in labels:
                        try:
                            if label.is_displayed():
                                driver.execute_script("arguments[0].click();", label)
                                print("   ✅ Label «информирование» нажат")
                                clicked = True
                                break
                        except Exception:
                            continue
                if not clicked:
                    print("   ⚠️  Не удалось включить информирование автоматически.")
                    print("   ⚠️  Задание будет создано — включите вручную в МТС если нужно.")
            except Exception as e:
                print(f"   ⚠️  Ошибка при включении информирования: {e}")
        print("   Продолжаю...")
        time.sleep(0.5)

        # Финальная проверка (только локально — в облаке некому смотреть)
        if not IS_CLOUD:
            print("\n" + "=" * 55)
            print("  ПРОВЕРЬ НА ЭКРАНЕ БРАУЗЕРА:")
            print("  - правильный текст сообщения?")
            print("  - правильное количество номеров?")
            print("  - расписание: Пн-Пт, 07:00-21:00?")
            print("  - максимум попыток: 3, перезвон через: 5 минут?")
            print("  - включён ли тумблер «Включить информирование»?")
            print("=" * 55)
            print("  Если всё верно — нажми Enter, чтобы сохранить и запустить обзвон.")
            print("  Если что-то не так — закрой окно браузера вместо Enter.")
            input("  -> Нажми Enter: ")
        else:
            print("☁️  Облачный режим: сохраняю задание автоматически...")

        # 12. Сохраняем
        print("💾 Сохраняю задание (это запускает обзвон)...")
        save_btn = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.XPATH, "//button[contains(., 'Сохранить')]"))
        )
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", save_btn)
        time.sleep(0.5)
        try:
            save_btn.click()
        except Exception:
            driver.execute_script("arguments[0].click();", save_btn)
        time.sleep(3)

        print(f"\n✅ Обзвон запущен!")
        print(f"   Ссылка на задание: {driver.current_url}")
        print(f"\n   Сохрани эту ссылку — она понадобится команде fetch_report.")
        print(f"   Обычно обзвон полностью завершается (с учётом")
        print(f"   перезвонов) примерно за 15 минут.")

        # Сохраняем ссылку в файл, чтобы START.bat мог подхватить её
        # автоматически в пункте «Полный цикл» — без ручного копирования.
        try:
            with open(LAST_TASK_URL_FILE, "w", encoding="utf-8") as f:
                f.write(driver.current_url.strip())
        except Exception as e:
            print(f"   ⚠️  Не удалось сохранить ссылку в файл {LAST_TASK_URL_FILE}: {e}")

        # Возвращаем url для использования в полном цикле
        return driver.current_url

    finally:
        if own_driver:
            driver.quit()
            print("Браузер закрыт.")



def cmd_fetch_report(task_url: str, external_driver=None, merge_fn=None, task_name_file=None, save_dir=None):
    """
    Команда: открыть список заданий МТС, показать последние задания,
    дать выбрать нужное, скачать отчёт и отправить результаты в реестр должников.

    merge_fn - какую функцию вызвать в конце для обработки отчёта
    (по умолчанию cmd_debtors_merge).
    task_name_file - где искать сохранённое название задания (по умолчанию
    DEBTOR_LAST_TASK_NAME_FILE).
    save_dir - куда сохранить локальную копию скачанного отчёта (по умолчанию
    папка output/).
    """
    if merge_fn is None:
        merge_fn = cmd_debtors_merge
    if task_name_file is None:
        task_name_file = DEBTOR_LAST_TASK_NAME_FILE
    if save_dir is None:
        save_dir = OUTPUT_DIR
    own_driver = external_driver is None
    driver = external_driver if external_driver else create_driver()
    try:
        print("🌐 Открываю список заданий МТС...")
        driver.get(MTS_TASKS_LIST_URL)
        time.sleep(2)
        logged_in_automatically = try_mts_login(driver)
        if logged_in_automatically:
            time.sleep(2)

        # Ждём загрузки таблицы заданий
        print("   Жду загрузки таблицы заданий МТС...")
        try:
            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "//tr[td]"))
            )
            time.sleep(1)
        except Exception:
            time.sleep(4)

        # Читаем сохранённое название задания
        task_name = None
        try:
            if os.path.exists(task_name_file):
                with open(task_name_file, encoding="utf-8") as f:
                    task_name = f.read().strip()
        except Exception:
            pass

        if not task_name:
            print("   ⚠️  Не найден файл с названием задания (last_task_name.txt).")
            if not IS_CLOUD:
                input("   Нажми кнопку скачивания вручную, затем Enter: ")
            selected_row = None
        else:
            print(f"   Ищу задание: «{task_name}»")

            # Ждём пока задание завершится (статус "Завершено" под названием)
            print("   Проверяю статус задания...")
            max_wait_minutes = 30
            for attempt in range(max_wait_minutes * 2):  # проверяем каждые 30 сек
                task_rows = driver.find_elements(By.XPATH, "//tr[td]")
                selected_row = None
                for tr in task_rows[:15]:
                    try:
                        name_el = tr.find_element(By.XPATH, ".//td[1]")
                        name = name_el.text.strip().split("\n")[0]
                        if task_name.lower() in name.lower() or name.lower() in task_name.lower():
                            full_text = name_el.text.strip().lower()
                            if "завершено" in full_text or "завершён" in full_text:
                                selected_row = tr
                                print(f"   ✅ Задание завершено: {name}")
                            else:
                                status = name_el.text.strip().split("\n")[1] if "\n" in name_el.text else "?"
                                print(f"   ⏳ Статус: «{status}» — жду завершения... ({attempt//2 + 1} мин)")
                            break
                    except Exception:
                        continue

                if selected_row is not None:
                    break

                if attempt < max_wait_minutes * 2 - 1:
                    time.sleep(30)
                    driver.refresh()
                    try:
                        WebDriverWait(driver, 15).until(
                            EC.presence_of_element_located((By.XPATH, "//tr[td]"))
                        )
                        time.sleep(1)
                    except Exception:
                        pass
            else:
                print(f"   ⚠️  Задание не завершилось за {max_wait_minutes} минут.")
                if IS_CLOUD:
                    print("   ⚠️  Облако: прерываю — запустите повторно позже.")
                    return
                print("   Скачай отчёт вручную когда задание завершится.")
                input("   -> Нажми Enter после скачивания: ")

        # Засекаем файлы ДО скачивания
        before_files = set(os.listdir(DOWNLOADS_DIR)) if os.path.isdir(DOWNLOADS_DIR) else set()

        download_clicked = False
        if selected_row:
            print("🔍 Ищу кнопку скачивания отчёта...")
            try:
                # Точный селектор кнопки скачивания (получен через DevTools):
                # div.itl-svg.svg-download-24.no-transition внутри td.itl-table-cell.report
                btn = WebDriverWait(selected_row, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR,
                        "div.svg-download-24, div[class*='svg-download'], .itl-svg.svg-download-24"
                    ))
                )
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
                time.sleep(0.5)
                driver.execute_script("arguments[0].click();", btn)
                download_clicked = True
                print("   ✅ Кнопка скачивания нажата")
            except Exception as e:
                print(f"   ⚠️  Не нашёл кнопку: {e}")

        if not download_clicked:
            print()
            print("   ⚠️  Не нашёл кнопку скачивания автоматически.")
            if IS_CLOUD:
                print("   ⚠️  Облако: прерываю — запустите повторно или скачайте вручную.")
                return
            print("   Сделай это вручную:")
            print("   1. Посмотри на открытый браузер — там список заданий МТС")
            print("   2. Нажми кнопку скачивания отчёта (иконка ↓) у нужного задания")
            print("   3. В появившемся окне выбери CSV и нажми «Скачать»")
            print("   4. Дождись скачивания и вернись сюда")
            input("   -> Нажми Enter после скачивания: ")

        # После клика по ↓ МТС показывает модальное окно выбора формата
        # (Excel / CSV) с кнопкой «Скачать» — всё на той же странице, без новой вкладки.
        # Выбираем CSV и жмём «Скачать».
        print("   Жду модальное окно выбора формата...")
        try:
            # Ждём появления кнопок выбора формата
            csv_btn = WebDriverWait(driver, 8).until(
                EC.element_to_be_clickable((By.XPATH,
                    "//button[normalize-space(.)='CSV'] | //label[normalize-space(.)='CSV']"
                ))
            )
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", csv_btn)
            time.sleep(0.2)
            csv_btn.click()
            print("   ✅ Выбран формат CSV")
            time.sleep(0.5)

            confirm_btn = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.XPATH,
                    "//button[contains(normalize-space(.),'Скачать')]"
                ))
            )
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", confirm_btn)
            time.sleep(0.2)
            confirm_btn.click()
            print("   ✅ Нажата кнопка «Скачать»")
        except Exception as e:
            print(f"   ⚠️  Модальное окно не появилось или формат уже выбран: {e}")

        # Ждём появления нового файла в Downloads
        # Важно: Chrome сначала создаёт файл с расширением .tmp (.crdownload),
        # и только после завершения переименовывает в настоящий файл.
        # Ждём пока появится готовый файл (не .tmp и не .crdownload).
        print("⏳ Жду скачивания файла...")
        new_file = None
        for _ in range(60):
            time.sleep(1)
            if os.path.isdir(DOWNLOADS_DIR):
                after_files = set(os.listdir(DOWNLOADS_DIR))
                diff = after_files - before_files
                # Игнорируем временные файлы Chrome
                diff = {f for f in diff if not f.endswith(".crdownload") and not f.endswith(".tmp")}
                if diff:
                    new_file = sorted(diff, key=lambda f: os.path.getmtime(
                        os.path.join(DOWNLOADS_DIR, f)))[-1]
                    break

        if not new_file:
            print("❌ Не удалось автоматически найти скачанный файл.")
            print(f"   Проверь папку {DOWNLOADS_DIR} вручную и запусти debtors_merge с этим файлом:")
            print(f"   python scrape_crm.py debtors_merge ИМЯ_ФАЙЛА")
            return

        report_path = os.path.join(DOWNLOADS_DIR, new_file)
        print(f"✅ Отчёт скачан: {report_path}")

        # Копируем отчёт в папку save_dir (для должников - в output/) для удобства
        os.makedirs(save_dir, exist_ok=True)
        local_copy = os.path.join(save_dir, os.path.basename(report_path))
        try:
            with open(report_path, "rb") as src, open(local_copy, "wb") as dst:
                dst.write(src.read())
            print(f"   Копия сохранена: {local_copy}")
            # Удаляем оригинал из Downloads — он уже скопирован в рабочую папку
            try:
                os.remove(report_path)
            except Exception:
                pass
            report_path = local_copy
        except Exception:
            pass

    finally:
        if own_driver:
            driver.quit()
            print("Браузер закрыт.")

    # Сразу запускаем объединение (с заказами CRM или с реестром должников)
    print("\n🔗 Запускаю объединение результатов...")
    merge_fn(report_path)



def cmd_debtors_full_cycle():
    """
    Полный цикл для ОБЗВОНА ДОЛЖНИКОВ: выгрузка номеров из реестра (7) +
    создание задания "Обзвон должников" в МТС (9) + скачивание отчёта и
    отправка результатов в реестр (8) - всё подряд, в одном браузере.
    """
    print("=" * 55)
    print("  ПОЛНЫЙ ЦИКЛ ОБЗВОНА ДОЛЖНИКОВ: шаг 1 из 3 - выгрузка из реестра")
    print("=" * 55)
    cmd_debtors_export()

    driver = create_driver()
    try:
        print()
        print("=" * 55)
        print("  ПОЛНЫЙ ЦИКЛ ОБЗВОНА ДОЛЖНИКОВ: шаги 2+3 - обзвон и отчёт")
        print("=" * 55)
        task_url = cmd_call_new_debtors(external_driver=driver)
        if not task_url:
            print("❌ Не удалось получить ссылку на задание, скачивание отчёта пропущено.")
            return
        cmd_fetch_report(task_url, external_driver=driver, merge_fn=cmd_debtors_merge)
    finally:
        driver.quit()
        print("Браузер закрыт.")

# ============================================================
#  ТОЧКА ВХОДА
# ============================================================

def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    command = sys.argv[1]

    if command == "debtors_export":
        cmd_debtors_export()
    elif command == "call_debtors":
        cmd_call_new_debtors()
    elif command == "fetch_debtors_report":
        task_url = sys.argv[2] if len(sys.argv) > 2 else MTS_TASKS_LIST_URL
        cmd_fetch_report(task_url, merge_fn=cmd_debtors_merge, task_name_file=DEBTOR_LAST_TASK_NAME_FILE, save_dir=OUTPUT_DIR)
    elif command == "debtors_full_cycle":
        cmd_debtors_full_cycle()
    elif command == "debtors_merge":
        if len(sys.argv) < 3:
            print("Укажи путь к файлу отчёта МТС:")
            print("  python scrape_crm.py debtors_merge ОТЧЕТ.csv")
            sys.exit(1)
        cmd_debtors_merge(sys.argv[2])
    else:
        print(f"Неизвестная команда: {command}")
        print(__doc__)
        sys.exit(1)


if __name__ == "__main__":
    main()
